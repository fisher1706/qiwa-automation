from pathlib import Path

from openpyxl import load_workbook


def get_value_from_xlsx_model(xlsm_file, sheet_of_page_translation, row, column):
    excel_file_path = Path(__file__).parent.parent.joinpath("data/dataportal").joinpath(xlsm_file)
    sheet = load_workbook(excel_file_path)[sheet_of_page_translation]
    cell_value = sheet.cell(row=row, column=column).value
    return cell_value
